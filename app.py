from flask import Flask, render_template, request
from openpyxl import load_workbook


app = Flask(__name__)
# Главная страница отображение ФОТО
@app.route("/")
def main():
    # Отображение на главной База данных xlsx
    excel = load_workbook("mybase.xlsx")
    page = excel["Фото"]
    images = []
    # Отображение по горизонтали ряды заголовок, URL, описание
    for row in page:
        titlee = row[0].value
        urll = row[1].value
        descriptionss = row[2].value
        lst_images = [row[0].value, row[1].value, row[2].value]
        images.append(lst_images)
    return render_template("main.html", images=images)

# Страница добавлния ФОТО [заголовок, URL, описание]
@app.route("/add", methods=["POST", "GET"])
def adds():
    if request.method == 'POST':
        title = request.form.get('title')
        urls = request.form.get('url')
        descriptions = request.form.get('description')
        # Запись в txt: База данных №1
        filee = open('uuu.txt', 'a+', encoding='utf-8')
        filee.write(str(urls) + ' '+ str(title) + '\n')
        filee.close()
        # Запись в xlsx лист Фото: База данных №2
        excel = load_workbook("mybase.xlsx")
        page = excel["Фото"]
        page.append([title, urls, descriptions])
        excel.save("mybase.xlsx")
    return render_template("photo.html")

    filee = open('uuu.txt', 'r', encoding='utf-8')
    filee.close()
    return render_template("photo.html")

#  Страница подробное отображения ФОТО: заголовок, url, описание в базы xlsx
@app.route("/page/<number>")
def page(number):
    excel = load_workbook("mybase.xlsx")
    page = excel["Фото"]
    lst_row = page[number] # чтение (отображение) ряда по горизонтали
    return render_template("page.html", lst = lst_row)